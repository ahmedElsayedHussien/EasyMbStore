import openpyxl
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from .models import Product, Stock, Warehouse
from .middleware import get_current_branch

@login_required
def download_import_template(request):
    # Create an empty Excel file with headers
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Define headers
    headers = [
        "الباركود (إجباري)", 
        "اسم الصنف (إجباري)", 
        "سعر البيع (إجباري)", 
        "متوسط التكلفة (اختياري)",
        "الكمية الافتتاحية (إجباري)",
        "النوع (phone, accessory, cover_screen, electrical, spare_part)"
    ]
    ws.append(headers)

    # Add sample row
    ws.append(["123456789", "جراب شفاف ايفون 13", 150, 100, 50, "accessory"])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="import_products_template.xlsx"'
    wb.save(response)
    return response

@login_required
def import_excel_data(request):
    branch = get_current_branch(request)
    # Default warehouse for the branch
    main_warehouse = Warehouse.objects.filter(branch=branch, is_active=True).first()

    if request.method == 'POST':
        if not main_warehouse:
            messages.error(request, "لم يتم العثور على مخزن نشط لهذا الفرع.")
            return redirect('erp:import_excel_data')

        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "يرجى اختيار ملف الإكسل.")
            return redirect('erp:import_excel_data')

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            success_count = 0
            skip_count = 0
            
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Unpack columns
                barcode = str(row[0]).strip() if row[0] else None
                name = str(row[1]).strip() if row[1] else None
                selling_price = row[2]
                cost = row[3] if row[3] else 0.0
                qty = row[4] if row[4] else 0
                product_type = str(row[5]).strip() if row[5] else 'accessory'
                
                # Validation
                if not barcode or not name or selling_price is None:
                    continue # Skip invalid rows
                    
                if barcode == 'None' or name == 'None':
                    continue

                # Check if product exists
                product, created = Product.objects.get_or_create(
                    barcode_qr=barcode,
                    branch=branch,
                    defaults={
                        'name': name,
                        'selling_price': selling_price,
                        'average_cost': cost,
                        'product_type': product_type,
                        'requires_imei': False # We assume bulk import is for non-serialized
                    }
                )

                # Add stock
                stock, _ = Stock.objects.get_or_create(
                    product=product,
                    warehouse=main_warehouse,
                    defaults={'quantity': 0}
                )
                
                try:
                    qty_int = int(float(qty))
                    if qty_int > 0:
                        stock.quantity += qty_int
                        stock.save()
                except ValueError:
                    pass

                if created:
                    success_count += 1
                else:
                    skip_count += 1

            messages.success(request, f"تم الاستيراد بنجاح! إضافة: {success_count} جديد، وتحديث: {skip_count} صنف.")
            return redirect('erp:setup_dashboard')
            
        except Exception as e:
            messages.error(request, f"حدث خطأ أثناء معالجة الملف: {str(e)}")
            return redirect('erp:import_excel_data')

    return render(request, 'erp/import_data.html')
